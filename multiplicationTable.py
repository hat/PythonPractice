__author__ = 'taztony2010'

#  This program is taken from Automate The Boring Stuff With Python by: Al Sweigart
#  Pg. 291 Practice Problem - Multiplication Table Maker
#
#  It creates a multiplication table of size (table_size) and writes it to an excel
#  sheet using openpyxl

import sys, openpyxl
from openpyxl.styles import Font, Style


#table_size = sys.argv[1]
table_size = 3
starting_num = 1
alphabet = ['A', 'B', 'C', 'D']

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

bold12style = Font(size = 12, bold = True)
styleObj = Style(font = bold12style)

print('Creating table...')

for columnNum in range(0, table_size):
    for rowNum in range(0, table_size):
        if columnNum == 0:
            sheet.cell(row = rowNum + 1, column = 1).font = bold12style
            sheet.cell(row = rowNum + 1, column = 1).value = rowNum + 1
        elif rowNum == 0:
            sheet.cell(row = 1, column = columnNum + 1).font = bold12style
            sheet.cell(row = 1, column = columnNum + 1).value = columnNum + 1
        else:
            currentCell = sheet.cell(row = rowNum + 1, column = columnNum + 1)
            currentCell.value = (columnNum + 1) * (rowNum + 1)


wb.save('multiplication_table_' + str(table_size) + '.xlsx')
print('Done.')